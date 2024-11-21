import paho.mqtt.client as mqtt
from datetime import datetime
from openpyxl import Workbook
import os

# Variables globales para almacenar los datos
dato1, dato2, dato3 = None, None, None

# Ruta al archivo Excel
file_path = "datos_estacion.xlsx"

# Función para crear el archivo Excel si no existe
def crear_excel():
    if not os.path.exists(file_path):
        # Crear un nuevo libro de trabajo y hoja de trabajo
        wb = Workbook()
        ws = wb.active
        # Agregar encabezados en la primera fila
        ws.append(["Fecha", "Hora", "LDR", "Temperatura", "Coordenada Motor"])
        wb.save(file_path)
        print(f"Archivo Excel creado: {file_path}")
    else:
        print(f"El archivo Excel ya existe: {file_path}")

# Función para guardar los datos en el archivo Excel
def guardar_datos_en_excel(fecha, hora, dato1, dato2, dato3):
    # Abrir el archivo existente para agregar datos
    from openpyxl import load_workbook
    wb = load_workbook(file_path)
    ws = wb.active
    # Agregar una nueva fila con los datos
    ws.append([fecha, hora, dato1, dato2, dato3])
    wb.save(file_path)  # Guardar los cambios
    print(f"Datos guardados en el archivo: {file_path}")

# Callback cuando se conecta al broker MQTT
def on_connect(client, userdata, flags, rc):
    print("Conectado al broker MQTT con código:", rc)
    client.subscribe("helion")  # Suscribirse al tema "helion"

# Callback cuando se recibe un mensaje del broker MQTT
def on_message(client, userdata, msg):
    global dato1, dato2, dato3
    
    # Decodifica el mensaje recibido
    mensaje = msg.payload.decode()
    print(f"Mensaje recibido de la estación: {msg.topic}: {mensaje}")
    
    # Divide el mensaje en los tres datos
    datos = mensaje.split(',')
    
    # Verifica que haya tres datos antes de asignarlos
    if len(datos) == 3:
        dato1 = datos[0]  # LDR
        dato2 = datos[1]  # Temperatura
        dato3 = datos[2]  # Coordenada del motor
        
        # Obtener la fecha y hora actuales
        fecha_hora_actual = datetime.now()
        fecha = fecha_hora_actual.strftime("%Y-%m-%d")
        hora = fecha_hora_actual.strftime("%H:%M:%S")
        
        # Imprimir los datos recibidos
        print(f"LDR Value: {dato1}")
        print(f"Temperatura: {dato2}° C")
        print(f"Coordenada en motor: {dato3}°")
        print(f"Fecha: {fecha}")
        print(f"Hora: {hora}")
        
        # Guardar los datos en el archivo Excel
        guardar_datos_en_excel(fecha, hora, dato1, dato2, dato3)

# Crear el archivo de Excel si no existe
crear_excel()

# Configuración del cliente MQTT
client = mqtt.Client()
client.on_connect = on_connect
client.on_message = on_message

# Conectarse al broker MQTT
client.connect("192.168.0.5", 1883, 60)

# Ejecutar el bucle principal de MQTT
client.loop_forever()
